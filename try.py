
#%% imports
import tempfile
import fitz
import shutil
import openpyxl
import os
import pandas as pd
from io import BytesIO
from datetime import datetime
import pytz
import psycopg2
from psycopg2 import Error
from fastapi import FastAPI, File, UploadFile

app = FastAPI()
def code_time_test_function():
    return "successfully tested code_time"
def another_code_time_test_function():
    return "successfullu testted code time another time"
def wakatime_test():
    return "at this point I test wakatime instead of code_time"

def decompose_tag(tag):
    if tag[:3] == "68-":
        start_pos=3
    else:
        start_pos=2
    pos_1 = tag.find("1")
    equip_cat=tag[start_pos:pos_1].replace("-","")
    unit = tag[pos_1:pos_1+2]
    if len(tag[pos_1:].replace("-",""))>5:
        package_letter = tag[pos_1+2]
        tag_number = tag[pos_1+3:pos_1+6]
        suffix = tag[pos_1+6:].replace("-","")
    else:
        package_letter = ""
        tag_number = tag[pos_1+2:pos_1+4]
        suffix = tag[pos_1+4:].replace("-","")
    return (equip_cat, unit, tag_number, suffix)

def create_or_overwrite_eqdb(connection, table_name, column_name, data_set):
    """
    Checks if a table exists. If it exists, it drops and recreates it 
    to effectively 'overwrite' the entire table structure and content. 
    It then inserts the unique values from the Python set.
    """
    
    # 1. Define the SQL to drop and recreate the table structure
    # Note: We use TEXT for the unique values, but this could be INTEGER if your set contains numbers.
    SQL_DROP = f"DROP TABLE IF EXISTS {table_name} CASCADE;"
    SQL_RECREATE = f"""
    CREATE TABLE {table_name} (
        id SERIAL PRIMARY KEY,
        {column_name} VARCHAR(40) UNIQUE NOT NULL
    );
    """
    SQL_INSERT_ROW = f"INSERT INTO {table_name} ({column_name}) VALUES (%s);"
    
    # Prepare data for bulk insertion: list of tuples (required by executemany)
    # The data set elements must be wrapped in a tuple: [(value,), (value,), ...]
    data_for_insert = [(item,) for item in data_set]
    
    try:
        with connection.cursor() as cursor:
            # --- Drop Existing Table (The Overwrite Step) ---
            cursor.execute(SQL_DROP)
            print(f"\nTable '{table_name}' dropped (if it existed) to prepare for overwrite.")
            
            # --- Create New Table ---
            cursor.execute(SQL_RECREATE)
            print(f"Table '{table_name}' successfully recreated.")
            
            # --- Bulk Insert Data ---
            if data_for_insert:
                cursor.executemany(SQL_INSERT_ROW, data_for_insert)
                connection.commit()
                print(f"Successfully inserted {len(data_for_insert)} unique rows into '{table_name}'.")
            else:
                print("The input set was empty; no data inserted.")
                
    except Error as e:
        print(f"Error during table creation/insertion: {e}")
        # Rollback the transaction in case of an error
        connection.rollback()        
    
eqdb_sheet_name = "NFE1-ME-20829-A-PO-F-001"


#%% connect to DB

DB_HOST = "localhost"
DB_NAME = "heru4_staging"
DB_USER = "python_service"
DB_PASSWORD = "08082018"
DB_PORT = "5432"

SQL_CREATE_DOC_TABLE = """
CREATE TABLE IF NOT EXISTS document_versions (
    doc_id VARCHAR(25) PRIMARY KEY,
    revision_number VARCHAR(2) NOT NULL
);
"""

SQL_CREATE_CLDT_TABLE = """
CREATE TABLE IF NOT EXISTS cldt (
    id SERIAL PRIMARY KEY,
    doc_id VARCHAR(25) NOT NULL,
    doc_part VARCHAR(3) NOT NULL,
    revision_number VARCHAR(2) NOT NULL,
    link_level VARCHAR(15) NOT NULL,
    tag VARCHAR(40) NOT NULL
);
"""

SQL_CREATE_ERRORS_TABLE = """
CREATE TABLE IF NOT EXISTS errors (
    id SERIAL PRIMARY KEY,
    doc_id VARCHAR(25) NOT NULL,
    revision_number VARCHAR(2) NOT NULL,
    page INTEGER NOT NULL,
    wrong_tag VARCHAR(40) NOT NULL,
    right_tag VARCHAR(40) NOT NULL
);
"""

#%% EQDB export to Postgres
@app.post("/update_eqdb/")
async def update_eqdb(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        EQDB_file_path = tmp.name
    workbook = openpyxl.load_workbook(EQDB_file_path,data_only=True)
    sheet = workbook[eqdb_sheet_name]

    eqdb_tags = set([cell[0].value for cell in sheet.iter_rows(11,sheet.max_row,2,2)])
    eqdb_tags.remove(None)

    try:
    # Establish the connection
        connection = psycopg2.connect(
                user=DB_USER,
                password=DB_PASSWORD,
                host=DB_HOST,
                port=DB_PORT,
                database=DB_NAME
                )

    # Cursor allows us to execute SQL commands
        cursor = connection.cursor()
        print("PostgreSQL database connection successful.")

        create_or_overwrite_eqdb(connection, "eqdb", "tag", eqdb_tags)

    except (Exception, Error) as error:
    # Catch connection and query errors
        print(f"Error while connecting to PostgreSQL or executing query: {error}")

    finally:
    # This block always executes, ensuring the connection is closed
        if connection:
            cursor.close()
            connection.close()
            print("\nPostgreSQL connection closed.")

def get_set_from_db(connection, table_name, column_name):
        # A. Define the SELECT query
    SQL_FETCH_CATEGORIES = f"SELECT {column_name} FROM {table_name};"
    
    try:
        with connection.cursor() as cursor:
            # B. Execute the query
            cursor.execute(SQL_FETCH_CATEGORIES)
            
            # C. Fetch all results (returns a list of tuples: [('electronics',), ('home',), ...])
            db_records = cursor.fetchall()
            
            # D. Transform the data into a set
            # This uses a generator expression inside the set() constructor for efficiency.
            # It iterates through the list of tuples (row), extracting the first element (column value) 
            # and feeding it directly to the set.
            eqdb_tags_set = {record[0] for record in db_records}
            
            return eqdb_tags_set
    except Error as e:
        print(f"Error fetching data to convert to set: {e}")
        connection.rollback()   
 

def ins_or_upd_doc_rev(connection, doc_id, new_revision):
    """
    Inserts a document revision. If the doc_id exists, 
    it updates the revision ONLY if the new_revision is greater than the existing one 
    based on the custom alphanumeric comparison rules (Alpha < Numeric).

    Returns True if the row was updated/inserted, False otherwise.
    """
    # SQL uses the ON CONFLICT clause targeting the PRIMARY KEY (doc_id)
    SQL_UPSERT_REVISION = """
    INSERT INTO document_versions (doc_id, revision_number)
    VALUES (%s, %s)
    ON CONFLICT (doc_id) 
    DO UPDATE SET 
        revision_number = EXCLUDED.revision_number 
    WHERE 
        -- --- Conditional Logic for Alphanumeric Comparison ---
        CASE 
            -- RULE 1: If CURRENT is Alpha AND NEW is Numeric, ALWAYS UPDATE (Numeric > Alpha)
            WHEN document_versions.revision_number ~ '^[A-Za-z]+$' AND EXCLUDED.revision_number ~ '^[0-9]+$' 
                THEN TRUE

            -- RULE 2: If CURRENT and NEW are BOTH Alpha (Compare alphabetically/lexicographically)
            WHEN document_versions.revision_number ~ '^[A-Za-z]+$' AND EXCLUDED.revision_number ~ '^[A-Za-z]+$' 
                THEN document_versions.revision_number < EXCLUDED.revision_number
                
            -- RULE 3: If CURRENT and NEW are BOTH Numeric (Compare numerically for correct natural sort: 2 < 10)
            -- We must safely cast to INTEGER for comparison.
            WHEN document_versions.revision_number ~ '^[0-9]+$' AND EXCLUDED.revision_number ~ '^[0-9]+$' 
                THEN document_versions.revision_number::INTEGER < EXCLUDED.revision_number::INTEGER
            
            -- RULE 4: If CURRENT is Numeric AND NEW is Alpha, NEVER UPDATE (Alpha < Numeric, so the existing numeric is higher)
            WHEN document_versions.revision_number ~ '^[0-9]+$' AND EXCLUDED.revision_number ~ '^[A-Za-z]+$' 
                THEN FALSE

            -- Default/Fallback: No update, or complex mixed alphanumeric strings not covered by above rules
            ELSE FALSE
        END;
    """
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(SQL_UPSERT_REVISION, (doc_id, new_revision))
            connection.commit()
            
            # Use rowcount to determine if any row was affected (inserted OR updated)
            return cursor.rowcount > 0
            
    except Error as e:
        print(f"Error during document revision upsert: {e}")
        connection.rollback()
        return False

 #%% file treatment

@app.post("/process_docs/")
async def process_documents(files: list[UploadFile] = File(...)):
#%% EQDB import from Postgres

    try:
        # Establish the connection
        connection = psycopg2.connect(
                user=DB_USER,
                password=DB_PASSWORD,
                host=DB_HOST,
                port=DB_PORT,
                database=DB_NAME
                )
        cursor = connection.cursor()
        eqdb_tags = get_set_from_db(connection, "eqdb", "tag")
    except (Exception, Error) as error:
            # Catch connection and query errors
        print(f"Error while connecting to PostgreSQL or executing query: {error}")

    finally:
        # This block always executes, ensuring the connection is closed
        if connection:
            cursor.close()
        connection.close()
        # print("\nPostgreSQL connection closed.")

    eqdb_dict = {decompose_tag(tag) : tag for tag in eqdb_tags}
    eqdb_decomposed = set([decompose_tag(tag) for tag in eqdb_tags])
    # --- 2. SQL Commands ---
    # Example: Create a new table

    #%%
    #%% documents register
    doc_reg = {}
    document_revisions = {}

    

    try:
        # Establish the connection
        connection = psycopg2.connect(
            user=DB_USER,
            password=DB_PASSWORD,
            host=DB_HOST,
            port=DB_PORT,
            database=DB_NAME
        )
        cursor = connection.cursor()
        cursor.execute(SQL_CREATE_DOC_TABLE)
        cursor.execute(SQL_CREATE_CLDT_TABLE)
        cursor.execute(SQL_CREATE_ERRORS_TABLE)
        connection.commit()
        for file in files:
            file_name = file.filename
            pdf_bytes = await file.read()
            await file.close()
            pdf_file = fitz.open(stream = pdf_bytes, filetype = "pdf")
            
            content_of_title_page = pdf_file[0].get_text("words",sort=False)
            doc_number_found = False
            revision_found = False
            for word in content_of_title_page:
                
                if ("3945_" in word[4]) and (len(word[4])==22):
                    document_number = word[4]
                    doc_number_found = True
                elif word[4]=="REV":
                    rev_x_pos = word[0]
                    rev_y_pos = word[1]
                    received_y_pos = rev_y_pos - 200
                    revision_found = True
            for word in content_of_title_page:
                if (word[0]> rev_x_pos-10) and (word[0] < rev_x_pos +10):
                    if (word[1]<rev_y_pos) and (word[1]>received_y_pos):
                        rev_y_pos = word[1]
                        document_revisions[document_number] = word[4]
            if not(doc_number_found and revision_found):
                document_number = file_name[:22]
                document_revisions[document_number] = file_name[-8:-6]
            treat_document = ins_or_upd_doc_rev(connection,
                                                document_number,
                                                document_revisions[document_number])
            if not(treat_document):
                continue
            #% tags extraction
            
            SQL_DELETE_PREVIOUS_TAGS = "DELETE FROM cldt WHERE doc_id = %s;"
            cursor.execute(SQL_DELETE_PREVIOUS_TAGS, (document_number,))
            connection.commit()
            
            tags_found = set()
            list_suspect = []
            page_num = 1
            for page in pdf_file:
                content_of_page = page.get_text("words",sort=False)
                if page_num == 1:
                    matrix = page.rotation_matrix
                for word in content_of_page:
                    if not(word[4] in tags_found):
                        word_decomposed = decompose_tag(word[4])
                        if (word[4] in eqdb_tags):
                            tags_found.add(word[4])
                        elif word_decomposed in eqdb_decomposed:
                            tags_found.add(eqdb_dict[word_decomposed])
                            list_suspect.append([document_number,
                                                 document_revisions[document_number],
                                                 page_num,
                                                 word[4] ,
                                                 eqdb_dict[word_decomposed]])
                        elif (len(word[4]) in {4, 5, 6}) and (word[4][:2]=="68"):
                            if page.rotation_matrix == matrix:
                                ending_coord = [word[0]-5,
                                                word[3],
                                                word[2]+5,
                                                2*word[3]-word[1]]                    
                            else:
                                ending_coord = [word[2],
                                                word[1]-5,
                                                2*word[2]-word[0],
                                                word[3]+5]                                
                            ending = page.get_textbox(ending_coord)
                            if ending[-2:]=="\n+":
                                ending = ending[:-2]
                            instrum_word = word[4]+ending
                            instrum_tag_decomposed=decompose_tag(instrum_word)
                            if instrum_word in eqdb_tags:
                                tags_found.add(instrum_word)
                            elif instrum_tag_decomposed in eqdb_decomposed:
                                tags_found.add(eqdb_dict[instrum_tag_decomposed])
                                list_suspect.append([document_number,
                                                     document_revisions[document_number],
                                                     page_num,
                                                     instrum_word ,
                                                     eqdb_dict[instrum_tag_decomposed]])                                   
                page_num += 1
            cldt_list = [[document_number,
                          "000",
                          document_revisions[document_number],
                          "Tag",item] for item in tags_found]
            SQL_INSERT_CLDT = """
            INSERT INTO cldt (doc_id, doc_part, revision_number, link_level, tag)
            VALUES (%s, %s, %s, %s, %s);"""
            cursor.executemany(SQL_INSERT_CLDT, cldt_list)
            connection.commit()
            
            SQL_INSERT_ERRORS = """
            INSERT INTO errors (doc_id, revision_number, page, wrong_tag, right_tag)
            VALUES (%s, %s, %s, %s, %s);"""
            cursor.executemany(SQL_INSERT_ERRORS, list_suspect)
            connection.commit()
            
            doc_reg[document_number] = tags_found
            pdf_file.close()
        
        
        
        # #%% convert doc register to list
        # cldt_list = []
        # for key in doc_reg.keys():
        #     for item in doc_reg[key]:
        #         cldt_list.append(["",key,"000",document_revisions[key],"Tag",item])
                
        # cldt_df = pd.DataFrame(cldt_list,columns = ["","Document","Level","revision","idk","Tag"])
        
        # list_suspect_df = pd.DataFrame(list_suspect,columns = ["document number","revision","page number", "wrong tag", "to be replaced with..."])


    except (Exception, Error) as error:
        # Catch connection and query errors
        print(f"Error while connecting to PostgreSQL or executing query: {error}")

    finally:
        # This block always executes, ensuring the connection is closed
        if connection:
            cursor.close()
            connection.close()
            print("\nPostgreSQL connection closed.")
#
