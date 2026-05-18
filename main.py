#%% imports
import config
import sql_commands as sql
import fitz
import openpyxl
import os
import pandas as pd
from datetime import datetime
import pytz
import psycopg2
from psycopg2 import Error

def decompose_tag(tag, tag_system_prefix=config.TAG_SYSTEM_PREFIX):
    if tag[:len(tag_system_prefix) + 1] == tag_system_prefix + "-":
        start_pos=len(tag_system_prefix) + 1
    else:
        start_pos=len(tag_system_prefix) 
    pos_1 = tag.find("1")
    equip_cat=tag[start_pos:pos_1].replace("-","")
    unit = tag[pos_1:pos_1+2]
    if len(tag[pos_1:].replace("-",""))>5:
        tag_number = tag[pos_1+3:pos_1+6]
        suffix = tag[pos_1+6:].replace("-","")
    else:
        tag_number = tag[pos_1+2:pos_1+4]
        suffix = tag[pos_1+4:].replace("-","")
    return (equip_cat, unit, tag_number, suffix)
    
def create_or_overwrite_eqdb(connection, table_name, column_name, data_set):
    """
    Checks if a table exists. If it exists, it drops and recreates it 
    to effectively 'overwrite' the entire table structure and content. 
    It then inserts the unique values from the Python set.
    """
    
    # 1. Define the SQL to drop and recreate the table structure
    # Note: We use TEXT for the unique values, but this could be INTEGER if your set contains numbers.
    SQL_DROP = f"DROP TABLE IF EXISTS {table_name} CASCADE;"
    SQL_RECREATE = f"""
    CREATE TABLE {table_name} (
        id SERIAL PRIMARY KEY,
        {column_name} VARCHAR(40) UNIQUE NOT NULL
    );
    """
    SQL_INSERT_ROW = f"INSERT INTO {table_name} ({column_name}) VALUES (%s);"
    
    # Prepare data for bulk insertion: list of tuples (required by executemany)
    # The data set elements must be wrapped in a tuple: [(value,), (value,), ...]
    data_for_insert = [(item,) for item in data_set]
    
    try:
        with connection.cursor() as cursor:
            # --- Drop Existing Table (The Overwrite Step) ---
            cursor.execute(SQL_DROP)
            print(f"\nTable '{table_name}' dropped (if it existed) to prepare for overwrite.")
            
            # --- Create New Table ---
            cursor.execute(SQL_RECREATE)
            print(f"Table '{table_name}' successfully recreated.")
            
            # --- Bulk Insert Data ---
            if data_for_insert:
                cursor.executemany(SQL_INSERT_ROW, data_for_insert)
                connection.commit()
                print(f"Successfully inserted {len(data_for_insert)} unique rows into '{table_name}'.")
            else:
                print("The input set was empty; no data inserted.")
                
    except Error as e:
        print(f"Error during table creation/insertion: {e}")
        # Rollback the transaction in case of an error
        connection.rollback()        
    
def get_set_from_db(connection, table_name, column_name):
        # A. Define the SELECT query
    SQL_FETCH_CATEGORIES = f"SELECT {column_name} FROM {table_name};"
    
    try:
        with connection.cursor() as cursor:
            # B. Execute the query
            cursor.execute(SQL_FETCH_CATEGORIES)
            
            # C. Fetch all results (returns a list of tuples: [('electronics',), ('home',), ...])
            db_records = cursor.fetchall()
            
            # D. Transform the data into a set
            # This uses a generator expression inside the set() constructor for efficiency.
            # It iterates through the list of tuples (row), extracting the first element (column value) 
            # and feeding it directly to the set.
            eqdb_tags_set = {record[0] for record in db_records}
            
            return eqdb_tags_set
    except Error as e:
        print(f"Error fetching data to convert to set: {e}")
        connection.rollback()   
        
def insert_or_update_document_revision(connection, doc_id, new_revision):
    """
    Inserts a document revision. If the doc_id exists, 
    it updates the revision ONLY if the new_revision is greater than the existing one 
    based on the custom alphanumeric comparison rules (Alpha < Numeric).

    Returns True if the row was updated/inserted, False otherwise.
    """
    # SQL uses the ON CONFLICT clause targeting the PRIMARY KEY (doc_id)
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql.SQL_UPSERT_REVISION, (doc_id, new_revision))
            connection.commit()
            
            # Use rowcount to determine if any row was affected (inserted OR updated)
            return cursor.rowcount > 0
            
    except Error as e:
        print(f"Error during document revision upsert: {e}")
        connection.rollback()
        return False
    

def eqdb_export_to_Postgres(connection):
    workbook = openpyxl.load_workbook(config.EQDB_file_path,data_only=True)
    sheet = workbook[config.eqdb_sheet_name]
    eqdb_tags = set([cell[0].value for cell in sheet.iter_rows(11,sheet.max_row,2,2)])
    eqdb_tags.remove(None)
    create_or_overwrite_eqdb(connection, "eqdb", "tag", eqdb_tags)
        
def eqdb_import(connection):
    global eqdb_tags, eqdb_dict, eqdb_decomposed
    eqdb_tags = get_set_from_db(connection, "eqdb", "tag")
    eqdb_dict = {decompose_tag(tag,config.TAG_SYSTEM_PREFIX) : tag for tag in eqdb_tags}
    eqdb_decomposed = set([decompose_tag(tag,config.TAG_SYSTEM_PREFIX) for tag in eqdb_tags])

def doc_num_and_rev(pdf_file, file_name):
    doc_number_found = False
    revision_found = False
    content_of_title_page = pdf_file[0].get_text("words",sort=False)
    for word in content_of_title_page:
        if (config.FILE_NUMBER_START in word[4]) and (len(word[4])==22):
            document_number = word[4]
            doc_number_found = True
        elif word[4]=="REV":
            rev_x_pos = word[0]
            rev_y_pos = word[1]
            received_y_pos = rev_y_pos - 200
            revision_found = True
            for word in content_of_title_page:
                if (word[0]> rev_x_pos-10) and (word[0] < rev_x_pos +10):
                    if (word[1]<rev_y_pos) and (word[1]>received_y_pos):
                        rev_y_pos = word[1]
                        doc_revision = word[4]
            if not(doc_number_found and revision_found):
                document_number = file_name[:22]
                doc_revision = file_name[-8:-6]
    return document_number, doc_revision



def get_files():
    all_entries = os.listdir(config.directory)
    # Filter the list to include only actual files
    only_files = [
                entry for entry in all_entries 
                if os.path.isfile(os.path.join(config.directory, entry))
                ]
    return only_files

def import_cldt(eqdb_decomposed):
    workbook = openpyxl.load_workbook(config.excel_file_path,data_only=True)
    sheet = workbook[config.cldt_sheet_name]
    imported_cldt = list(sheet.iter_rows(7,sheet.max_row,2,6, values_only=True))
    imported_cldt = [row for row in imported_cldt if decompose_tag(row[4],config.TAG_SYSTEM_PREFIX) in eqdb_decomposed]
    return imported_cldt

def analyze_file(pdf_file, document_number, doc_rev, eqdb_tags, eqdb_dict):
    tags_found = set()
    list_suspect = []
    page_num = 1
    matrix = pdf_file[1].rotation_matrix
    for page in pdf_file:
        content_of_page = page.get_text("words",sort=False)
        for word in content_of_page:
            if not(word[4] in tags_found):
                word_decomposed = decompose_tag(word[4],config.TAG_SYSTEM_PREFIX)
                if (word[4] in eqdb_tags):
                    tags_found.add(word[4])
                elif word_decomposed in eqdb_decomposed:
                    tags_found.add(eqdb_dict[word_decomposed])
                    list_suspect.append([document_number, doc_rev, page_num,
                                         word[4] , eqdb_dict[word_decomposed]])
                elif (len(word[4]) in {4, 5, 6}) and (word[4][:2]==config.TAG_SYSTEM_PREFIX):
                    if page.rotation_matrix == matrix:
                        ending_coord = [word[0]-5,word[3],word[2]+5,2*word[3]-word[1]]                    
                    else:
                        ending_coord = [word[2],word[1]-5,2*word[2]-word[0],word[3]+5]                                
                    ending = page.get_textbox(ending_coord)
                    if ending[-2:]=="\n+":
                        ending = ending[:-2]
                    instrum_word = word[4]+ending
                    instrum_tag_decomposed=decompose_tag(instrum_word,config.TAG_SYSTEM_PREFIX)
                    if instrum_word in eqdb_tags:
                        tags_found.add(instrum_word)
                    elif instrum_tag_decomposed in eqdb_decomposed:
                        tags_found.add(eqdb_dict[instrum_tag_decomposed])
                        list_suspect.append([document_number, doc_rev,
                                             page_num,  instrum_word ,
                                             eqdb_dict[instrum_tag_decomposed]])                                   
        page_num += 1
    return tags_found, list_suspect

def merge_imported(imported_cldt, tags_found, document_number):
    tags_namrata = [row[4] for row in imported_cldt if row[0] == document_number]
    tags_found.update(tags_namrata)
    return tags_found

def cldt_directly(connection, cursor, imported_cldt):
    treated_files = get_set_from_db(connection,
                                    "document_versions",
                                    "doc_id")
    cldt_list = []
    for i in range(len(imported_cldt)):
        doc_num = imported_cldt[i][0] 
        if (doc_num not in treated_files):
            if (doc_num != imported_cldt[i-1][0]):
                update = insert_or_update_document_revision(connection, doc_num, imported_cldt[i][2])
            cldt_list.append(imported_cldt[i])
    cursor.executemany(sql.SQL_INSERT_CLDT, cldt_list)
    connection.commit()

def file_treatment(connection, cursor, imported_cldt):
    cursor.execute(sql.SQL_CREATE_DOC_TABLE)
    cursor.execute(sql.SQL_CREATE_CLDT_TABLE)
    cursor.execute(sql.SQL_CREATE_ERRORS_TABLE)
    connection.commit()
    document_revisions = {}
    only_files = get_files() 
    for file_name in only_files:
        file_path = os.path.join(config.directory,file_name)

        pdf_file = fitz.open(file_path)
        document_number, document_revisions[document_number] = \
                doc_num_and_rev(pdf_file, file_name)
        treat_document = insert_or_update_document_revision(
                connection, document_number,
                document_revisions[document_number])
        if not(treat_document):
            continue
        cursor.execute(sql.SQL_DELETE_PREVIOUS_TAGS, (document_number,))
        connection.commit()
        tags_found, list_suspect = analyze_file(pdf_file, document_number,
                                                document_revisions[document_number],
                                                eqdb_tags,eqdb_dict)
        tags_found = merge_imported(imported_cldt, tags_found, document_number)
        cldt_list = [[document_number, "000", document_revisions[document_number],"Tag",item] for item in tags_found]
        cursor.executemany(sql.SQL_INSERT_CLDT, cldt_list)
        connection.commit()
        cursor.executemany(sql.SQL_INSERT_ERRORS, list_suspect)
        connection.commit()
    



def check_remaining(connection):

    workbook = openpyxl.load_workbook(config.vdl_file_path,data_only=True)
    sheet = workbook[config.vdl_sheet_name]
    total_docs = set([cell[0] for cell in sheet.iter_rows(6,sheet.max_row,2,2, values_only=True)])
    files_complete_list = pd.read_sql(f"SELECT * FROM document_versions;",connection).values.tolist()
    treated_docs = set([row[0] for row in files_complete_list])
    diff = total_docs - treated_docs
    df = pd.DataFrame(list(diff))
    df.to_excel("not_treated.xlsx", index=False)



def export_table_to_excel(connection, table_name, output_file):
    """
    Connects to PostgreSQL, fetches an entire table into a DataFrame,
    and saves it to an Excel file.
    """
    # Construct the SQL query
    sql_query = f"SELECT * FROM {table_name};"
    # Use pandas.read_sql to execute the query and load results directly into a DataFrame
    df = pd.read_sql(sql_query, connection)
    # Save the DataFrame to an Excel file
    # index=False prevents writing the DataFrame's numerical index to the file
    df.to_excel(output_file, index=False)
    print(f"\nSuccessfully exported table '{table_name}' to '{output_file}'.")

def export_tables(connection):
    TABLES_TO_EXPORT = ["cldt", "errors", "document_versions"]
    timestamp = datetime.now(pytz.timezone('Europe/Paris')).strftime("%Y-%m-%d_%H_%M_%S")
    for table in TABLES_TO_EXPORT:
        OUTPUT_EXCEL_FILE = f"{table}_export_{timestamp}.xlsx"
        export_table_to_excel(connection, table, OUTPUT_EXCEL_FILE)

def export_untreated_files(connection):
    # Use pandas.read_sql to execute the query and load results directly into a DataFrame
    cldt_complete_list = pd.read_sql(f"SELECT * FROM cldt;",
                                     connection).values.tolist()
    files_complete_list = pd.read_sql(f"SELECT * FROM document_versions;",
                                      connection).values.tolist()
    documents_w_tags = set([row[1] for row in cldt_complete_list])
    treated_docs = set([row[0] for row in files_complete_list])
    docs_wo_tags = []
    for doc in treated_docs:
        if doc not in documents_w_tags:
            docs_wo_tags.append(doc)
    df_docs_wo_tags = pd.DataFrame(docs_wo_tags)
    df_docs_wo_tags.to_excel("docs_no_tags.xlsx", index = False)


def delete_tables(connection, cursor):

    for table_name in config.tables_to_delete:
        SQL_DROP = f"DROP TABLE IF EXISTS {table_name} CASCADE;"
        cursor.execute(SQL_DROP)    
        connection.commit()
        print("%s deleted successfuly;" % (table_name,))

def main():
    """
    Main pipeline controller: orchestrates the entire tag extraction workflow.
    
    Steps:
    1. Connect to database
    2. Optionally export EQDB to Postgres (if BOOL_EXPORT_EQDB is True)
    3. Optionally delete existing tables (if BOOL_DELETE_TABLES is True)
    4. Import EQDB tags from database
    5. Process PDF files and extract tags
    6. Check remaining untreated documents
    7. Export results to Excel files
    8. Export list of untreated files
    
    All database connections are properly managed and cleaned up in finally block.
    """
    connection = None
    cursor = None
    
    try:
        # Step 1: Connect to the database
        connection = psycopg2.connect(
            user=config.DB_USER,
            password=config.DB_PASSWORD,
            host=config.DB_HOST,
            port=config.DB_PORT,
            database=config.DB_NAME
        )
        cursor = connection.cursor()
        print("Database connection successful.")
        
        # Step 2: Optionally export EQDB to Postgres
        if config.BOOL_EXPORT_EQDB:
            eqdb_export_to_Postgres(connection)
        
        # Step 3: Optionally delete existing tables (destructive operation)
        if config.BOOL_DELETE_TABLES:
            delete_tables(connection, cursor)
        
        # Step 4: Import EQDB tags from database
        eqdb_import(connection)
        
        # Step 5: Process PDF files and extract tags

        imported_cldt = import_cldt(eqdb_decomposed)
        file_treatment(connection, cursor, imported_cldt)

        cldt_directly(connection, cursor, imported_cldt)
        
        # Step 6: Check remaining untreated documents
        check_remaining(connection)
        
        # Step 7: Export results to Excel files
        export_tables(connection)
        
        # Step 8: Export list of untreated files
        export_untreated_files(connection)
        
        print("\n--- Pipeline execution complete ---")

    except (Exception, Error) as error:
        # Catch connection and query errors
        print(f"Error while connecting to PostgreSQL or executing query: {error}")
        # Re-raise to allow caller to handle if needed
        raise

    finally:
        # This block always executes, ensuring the connection is closed
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("\nPostgreSQL connection closed.")


if __name__ == "__main__":
    main()    
